# AutomacaoSelic/excel_updater.py

import pandas as pd
import logging
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def atualizar_todas_planilhas(pasta_alvo: Path, df_selic: pd.DataFrame, nome_aba: str):
    logging.info(f"Procurando planilhas em: '{pasta_alvo}'")

    arquivos = list(pasta_alvo.glob("*.xlsx")) + list(pasta_alvo.glob("*.xlsm"))

    if not arquivos:
        logging.warning("Nenhum arquivo .xlsx ou .xlsm encontrado na pasta.")
        return

    logging.info(f"Encontrados {len(arquivos)} arquivos. Iniciando atualização...")

    for arquivo in arquivos:
        try:
            if arquivo.suffix.lower() == ".xlsm":
                book = load_workbook(arquivo, keep_vba=True)
                logging.info(f"Processando '{arquivo.name}' (com macros)...")
            else:
                book = load_workbook(arquivo)
                logging.info(f"Processando '{arquivo.name}' (padrão)...")

            if nome_aba in book.sheetnames:
                aba_antiga = book[nome_aba]
                book.remove(aba_antiga)

            nova_aba = book.create_sheet(nome_aba)

            rows = dataframe_to_rows(df_selic, index=False, header=True)

            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    nova_aba.cell(row=r_idx + 2, column=c_idx + 1, value=value)

            formato_data = "dd/mm/yyyy"
            formato_numero = "0.00"

            for linha in range(4, nova_aba.max_row + 1):
                nova_aba.cell(row=linha, column=2).number_format = formato_data
                nova_aba.cell(row=linha, column=3).number_format = formato_numero

            book.save(arquivo)

            logging.info(f"Arquivo '{arquivo.name}' atualizado com sucesso.")

        except Exception as e:
            logging.error(f"FALHA ao processar o arquivo '{arquivo.name}': {e}")
