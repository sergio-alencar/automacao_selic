import pandas as pd
import logging
from pathlib import Path
from typing import List, Any, Tuple
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from config.constants import ExcelFormats
from excel.utils import get_excel_files


ROW_OFFSET: int = 2
COL_OFFSET: int = 1

FORMAT_START_ROW: int = 4
DATE_COL_INDEX: int = 2
NUMBER_COL_INDEX: int = 3


def _write_data_to_sheet(sheet: Any, df: pd.DataFrame) -> None:
    rows = dataframe_to_rows(df, index=False, header=True)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(
                row=r_idx + ROW_OFFSET,
                column=c_idx + COL_OFFSET,
                value=value,
            )


def _apply_sheet_formatting(sheet: Any) -> None:
    for row_idx in range(FORMAT_START_ROW, sheet.max_row + 1):
        sheet.cell(row=row_idx, column=DATE_COL_INDEX).number_format = ExcelFormats.DATE
        sheet.cell(row=row_idx, column=NUMBER_COL_INDEX).number_format = ExcelFormats.NUMBER


def _update_single_workbook(
    file_path: Path,
    selic_df: pd.DataFrame,
    sheet_name: str,
) -> None:
    keep_vba: bool = file_path.suffix.lower() == ".xlsm"
    book = load_workbook(file_path, keep_vba=keep_vba)

    logging.info(f"Processing '{file_path.name}'...")

    if sheet_name in book.sheetnames:
        book.remove(book[sheet_name])

    new_sheet = book.create_sheet(sheet_name)

    _write_data_to_sheet(new_sheet, selic_df)
    _apply_sheet_formatting(new_sheet)

    book.save(file_path)
    book.close()


def update_all_worksheets(
    target_files: List[Path],
    selic_df: pd.DataFrame,
    sheet_name: str,
) -> None:
    if not target_files:
        logging.warning("No .xlsx or .xlsm files provided for update.")
        return

    logging.info(f"Starting update for {len(target_files)}...")

    for file_path in target_files:
        try:
            _update_single_workbook(
                file_path=file_path,
                selic_df=selic_df,
                sheet_name=sheet_name,
            )
            logging.info(f"File '{file_path.name}' updated successfully.")

        except Exception as e:
            logging.error(f"FAILED to process the file '{file_path.name}': {e}")
